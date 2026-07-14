import pandas as pd
from openpyxl import load_workbook


def load_data(file_path) -> dict:
    """
    Load all sheets from an Excel file.
    Returns dict of {sheet_name: DataFrame}, skipping title/header rows
    that are not real data.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    result = {}
    for sheet in sheet_names:
        if sheet.lower() in ("summary dashboard", "summary"):
            continue  # skip computed summary sheets
        try:
            df = pd.read_excel(file_path, sheet_name=sheet, header=None, dtype=str)
            # Find the real header row: first row where first cell looks like a column name
            header_row = 0
            for i, row in df.iterrows():
                first_val = str(row.iloc[0]).strip()
                if first_val.startswith("PT-") or first_val == "nan":
                    break
                if first_val not in ("nan", "") and not any(
                    emoji in first_val for emoji in ["🏥","🩺","🔬","📊"]
                ):
                    header_row = i
                    break
            df = pd.read_excel(file_path, sheet_name=sheet, header=header_row)
            df = df.dropna(how="all")
            df.columns = [str(c).strip() for c in df.columns]
            # Drop rows where Patient ID is NaN or doesn't start with PT-
            if "Patient ID" in df.columns:
                df = df[df["Patient ID"].astype(str).str.match(r"PT-\d{4}")]
            result[sheet] = df.reset_index(drop=True)
        except Exception as e:
            result[sheet] = pd.DataFrame()  # empty on error
    return result


def get_quick_stats(df_dict: dict) -> dict:
    """Return quick stats for the sidebar."""
    stats = {}
    for sheet, df in df_dict.items():
        if df is not None and not df.empty:
            stats[sheet] = {
                "rows": len(df),
                "cols": len(df.columns),
                "columns": df.columns.tolist()
            }
    return stats
